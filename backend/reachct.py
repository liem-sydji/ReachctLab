"""
ReachCT — reachct.py
Main scraper. Searches Google Maps and extracts company contact info.

Requirements:
    pip install playwright openpyxl beautifulsoup4
    playwright install chromium

Usage:
    python reachct.py --query "Marketing Company" --city "Madrid" --country "Spain" --start 0 --end 25
"""

import re
import uuid
import random
import argparse
import asyncio
from urllib.parse import urlparse
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from database import init_db, save_search, upsert_company, get_companies


# ── CONFIG ────────────────────────────────────────────────────────────────────
HEADLESS = True

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
]

CONTACT_PATHS = [
    "/contact", "/contact-us", "/contacto",
    "/about", "/about-us",
]

EMAIL_BLACKLIST = [
    "example", "domain", "youremail", "user@", "email@",
    "sentry", "wix", "wordpress", "jquery", "schema",
    "png", "jpg", "gif", "svg", "css", "js@", "noreply",
]

MAX_RETRIES = 1  # 1 attempt per site
# ─────────────────────────────────────────────────────────────────────────────


# ── Helpers ───────────────────────────────────────────────────────────────────

def random_delay(min_ms=1000, max_ms=2500):
    return random.randint(min_ms, max_ms)


def extract_emails(text: str) -> list:
    found = re.findall(r"<?([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})>?", text)
    return list(dict.fromkeys([
        e.lower() for e in found
        if not any(b in e.lower() for b in EMAIL_BLACKLIST)
    ]))


def extract_phones(text: str) -> list:
    found = re.findall(r"(?:\+?\d[\s\-.]?){7,15}", text)
    return list(dict.fromkeys([
        p.strip() for p in found
        if 7 <= len(re.sub(r"\D", "", p)) <= 15
    ]))


# ── Website scraping ──────────────────────────────────────────────────────────

async def try_contact_pages(page, base_url: str) -> dict:
    parsed = urlparse(base_url)
    base   = f"{parsed.scheme}://{parsed.netloc}"
    for path in CONTACT_PATHS:
        try:
            r = await page.goto(base + path, timeout=6000, wait_until="domcontentloaded")
            if r and r.status == 200:
                await page.wait_for_timeout(random_delay(800, 1500))
                html   = await page.content()
                text   = await page.evaluate("() => document.body.innerText")
                emails = extract_emails(html + " " + text)
                if emails:
                    phones = extract_phones(html + " " + text)
                    return {"emails": emails, "phones": phones}
        except:
            continue
    return {"emails": [], "phones": []}


async def scrape_website(browser, url: str, retries: int = MAX_RETRIES) -> dict:
    if not url:
        return {"email": "", "phone": "", "page_text": ""}

    for attempt in range(retries):
        context = None
        try:
            context   = await browser.new_context(user_agent=random.choice(USER_AGENTS))
            page      = await context.new_page()
            email     = ""
            phone     = ""
            page_text = ""

            await page.goto(url, timeout=15000, wait_until="domcontentloaded")
            await page.wait_for_timeout(random_delay(1500, 2500))

            try:
                html      = await page.content()
                page_text = await page.evaluate("() => document.body.innerText")
                if len(page_text.strip()) < 100:
                    await page.wait_for_timeout(2000)
                    html      = await page.content()
                    page_text = await page.evaluate("() => document.body.innerText")
            except:
                html      = ""
                page_text = ""

            emails = extract_emails(html + " " + page_text)
            phones = extract_phones(html + " " + page_text)

            if emails: email = emails[0]
            if phones: phone = phones[0]

            if not email:
                data = await try_contact_pages(page, url)
                if data["emails"]: email = data["emails"][0]
                if not phone and data["phones"]: phone = data["phones"][0]

            return {"email": email, "phone": phone, "page_text": page_text}

        except Exception as e:
            print(f"    ⚠️  Failed: {str(e)[:60]}")
        finally:
            if context:
                try: await context.close()
                except: pass

    return {"email": "", "phone": "", "page_text": ""}


# ── Name extraction ───────────────────────────────────────────────────────────

async def get_business_name(page) -> str:
    SKIP = {"resultados", "results", "google maps", "google", "", "maps"}

    for panel_sel in ['div[role="main"]', 'div.m6QErb']:
        try:
            await page.wait_for_selector(panel_sel, timeout=4000)
            break
        except:
            continue

    await page.wait_for_timeout(random_delay(600, 1000))

    for sel in ['h1.DUwDvf', 'div.DUwDvf', 'h1.fontHeadlineLarge',
                'div.fontHeadlineLarge', 'div[role="main"] h1']:
        try:
            els = await page.locator(sel).all()
            for el in els:
                try:
                    candidate = (await el.inner_text(timeout=2000)).strip()
                    if candidate.lower() not in SKIP and len(candidate) > 1:
                        return candidate
                except:
                    continue
        except:
            continue

    try:
        slug = page.url.split("/maps/place/")[1].split("/")[0]
        name = slug.replace("+", " ").replace("%20", " ").strip()
        if name.lower() not in SKIP:
            return name
    except:
        pass

    return ""


# ── Google Maps scraper ───────────────────────────────────────────────────────

async def scrape_google_maps(query: str, city: str, country: str,
                              start_idx: int, end_idx: int, run_id: str,
                              jobs: dict = None, job_id: str = None) -> list:
    results  = []
    location = f"{city}, {country}"
    search   = f"{query} {location}"
    maps_url = f"https://www.google.com/maps/search/{search.replace(' ', '+')}"

    print(f"\n🗺️  ReachCT — Searching: '{search}'")
    print(f"   Range: {start_idx} → {end_idx}")
    print(f"   {maps_url}\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            locale="es-ES",
            viewport={"width": 1280, "height": 800},
        )
        page = await context.new_page()
        await page.goto(maps_url, timeout=30000)
        await page.wait_for_timeout(random_delay(1500, 2000))

        # Accept cookies
        try:
            for btn_text in ["Aceptar todo", "Accept all", "Aceptar", "Accept"]:
                btn = page.get_by_role("button", name=btn_text)
                if await btn.count() > 0:
                    await btn.first.click()
                    print("🍪 Accepted cookies")
                    await page.wait_for_timeout(random_delay(800, 1200))
                    break
        except:
            pass

        print(f"📄 {await page.title()}")
        print(f"📜 Scrolling to load listings {start_idx}→{end_idx}...")

        seen_hrefs   = set()
        all_hrefs    = []
        scroll_count = 0
        max_scrolls  = (end_idx // 3) + 6

        while len(all_hrefs) < end_idx and scroll_count < max_scrolls:
            links = await page.locator('a[href*="/maps/place/"]').all()
            for l in links:
                try:
                    href = await l.get_attribute("href") or ""
                    if "/maps/place/" in href and href not in seen_hrefs:
                        seen_hrefs.add(href)
                        all_hrefs.append(href)
                except:
                    continue

            if len(all_hrefs) >= end_idx:
                print(f"   ✅ Loaded {len(all_hrefs)} listings — stopping scroll")
                break

            try:
                for sel in ['div[role="feed"]', 'div[aria-label*="Resultados"]', 'div[aria-label*="Results"]']:
                    feed = page.locator(sel)
                    if await feed.count() > 0:
                        await feed.evaluate("el => el.scrollTop += 1500")
                        break
                await page.wait_for_timeout(random_delay(600, 1000))
            except:
                break
            scroll_count += 1

        batch         = all_hrefs[start_idx:end_idx]
        total         = len(batch)
        total_on_maps = len(all_hrefs)

        if jobs and job_id:
            jobs[job_id]["total_on_maps"] = total_on_maps
            jobs[job_id]["processing"]    = total

        print(f"\n✅ {total_on_maps} total listings — processing {start_idx}→{start_idx+total}\n")

        if total == 0:
            print("❌ No listings in range.")
            await browser.close()
            return []

        for i, listing_href in enumerate(batch):
            # Check for cancellation — hard stop, kill browser immediately
            if jobs and job_id and jobs.get(job_id, {}).get("status") == "cancelling":
                print(f"  🛑 Search hard-stopped at listing {start_idx+i+1} — saving {len(results)} results")
                try:
                    await browser.close()
                except:
                    pass
                return results

            # Restart browser every 10 listings to free RAM
            if i > 0 and i % 10 == 0:

                try: await browser.close()
                except: pass
                await asyncio.sleep(2)
                browser = await p.chromium.launch(headless=HEADLESS)
                context = await browser.new_context(
                    user_agent=random.choice(USER_AGENTS),
                    locale="es-ES",
                    viewport={"width": 1280, "height": 800},
                )
                page = await context.new_page()
                await page.goto(maps_url, timeout=15000, wait_until="domcontentloaded")
                await page.wait_for_timeout(random_delay(1500, 2000))
                try:
                    for btn_text in ["Aceptar todo", "Accept all", "Aceptar", "Accept"]:
                        btn = page.get_by_role("button", name=btn_text)
                        if await btn.count() > 0:
                            await btn.first.click()
                            await page.wait_for_timeout(800)
                            break
                except:
                    pass

            try:
                await page.goto(listing_href, timeout=15000, wait_until="domcontentloaded")
                await page.wait_for_timeout(random_delay(1500, 2000))

                name = await get_business_name(page)

                # Phone
                phone_maps = ""
                try:
                    btn = page.locator('button[data-item-id*="phone"]')
                    if await btn.count() > 0:
                        raw        = await btn.first.get_attribute("data-item-id") or ""
                        phone_maps = raw.replace("phone:tel:", "").strip()
                except:
                    pass

                # Website
                website = ""
                for web_sel in ['a[data-item-id="authority"]',
                                'a[aria-label*="sitio web" i]',
                                'a[aria-label*="website" i]']:
                    try:
                        el = page.locator(web_sel).first
                        if await el.count() > 0:
                            web_href = await el.get_attribute("href") or ""
                            if web_href and "google" not in web_href:
                                website = web_href
                                break
                    except:
                        continue

                print(f"[{start_idx+i+1}] {name or '(no name)'}")
                print(f"  📞 {phone_maps or 'no phone'}")
                print(f"  🌐 {website[:65] if website else 'no website'}")

                web_data = {"email": "", "phone": "", "page_text": ""}
                if website:
                    web_data = await scrape_website(browser, website)

                final_phone = phone_maps or web_data["phone"]
                email       = web_data["email"]
                print(f"  ✉️  {email or 'no email'}\n")

                results.append({
                    "run_id":       run_id,
                    "name":         name,
                    "email":        email,
                    "phone":        final_phone,
                    "website":      website,
                    "city":         city,
                    "country":      country,
                    "company_type": query,
                    "maps_url":     page.url,
                })

                await page.wait_for_timeout(random_delay(300, 600))

            except Exception as e:
                error_msg = str(e)
                print(f"  ⚠️  Error on listing {start_idx+i+1}: {error_msg[:80]}\n")
                if "Page crashed" in error_msg or "Target closed" in error_msg:

                    try: await browser.close()
                    except: pass
                    await asyncio.sleep(2)
                    browser = await p.chromium.launch(headless=HEADLESS)
                    context = await browser.new_context(
                        user_agent=random.choice(USER_AGENTS),
                        locale="es-ES",
                        viewport={"width": 1280, "height": 800},
                    )
                    page = await context.new_page()
                    await page.goto(maps_url, timeout=15000, wait_until="domcontentloaded")
                    await page.wait_for_timeout(random_delay(1500, 2000))

                continue

        await browser.close()

    return results


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(data: list, query: str, city: str, country: str) -> str:
    if not data:
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    hdr_fill  = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
    hdr_font  = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    alt_fill  = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    wht_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ttl_fill  = PatternFill(start_color="E8005A", end_color="E8005A", fill_type="solid")
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")
    d_font    = Font(name="Arial", size=10)
    thin      = Side(style="thin", color="EEEEEE")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["Company Name", "Email", "Phone Number", "Website", "Location", "Company Type"]
    col_widths = [32, 35, 20, 40, 22, 25]
    num_cols   = len(headers)

    ws.merge_cells(f"A1:{chr(64+num_cols)}1")
    t           = ws["A1"]
    t.value     = f"ReachCT  |  {query}  |  {city}, {country}  |  {datetime.now().strftime('%Y-%m-%d')}"
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.alignment = center
    t.fill      = ttl_fill
    ws.row_dimensions[1].height = 30

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 26

    for row_i, item in enumerate(data, 3):
        fill   = alt_fill if row_i % 2 == 0 else wht_fill
        values = [
            item.get("name",         ""),
            item.get("email",        ""),
            item.get("phone",        ""),
            item.get("website",      ""),
            f"{item.get('city','')}, {item.get('country','')}",
            item.get("company_type", ""),
        ]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row_i, column=col, value=val)
            cell.fill      = fill
            cell.font      = d_font
            cell.border    = border
            cell.alignment = left
        ws.row_dimensions[row_i].height = 20

    fr  = len(data) + 3
    ws.merge_cells(f"A{fr}:{chr(64+num_cols)}{fr}")
    f   = ws[f"A{fr}"]
    f.value = (
        f"Total: {len(data)}   |   "
        f"Emails: {sum(1 for r in data if r.get('email'))}   |   "
        f"Phones: {sum(1 for r in data if r.get('phone'))}   |   "
        f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    f.font      = Font(name="Arial", size=9, italic=True, color="888888")
    f.alignment = center
    ws.row_dimensions[fr].height = 18
    ws.freeze_panes = "A3"

    stamp    = datetime.now().strftime("%Y%m%d_%H%M")
    loc_slug = f"{city}_{country}".replace(" ", "_").replace(",", "")
    filename = f"reachct_{query.replace(' ','_')}_{loc_slug}_{stamp}.xlsx"
    wb.save(filename)
    print(f"\n📊 Excel saved: {filename}")
    return filename


# ── Entry point ───────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(description="ReachCT — Google Maps Contact Scraper")
    parser.add_argument("--query",   default="Marketing Company", help="Type of business")
    parser.add_argument("--city",    default="Madrid",  help="City to search in")
    parser.add_argument("--country", default="Spain",   help="Country to search in")
    parser.add_argument("--start",   default=0,  type=int)
    parser.add_argument("--end",     default=25, type=int)
    parser.add_argument("--export",  action="store_true")
    args = parser.parse_args()

    init_db()

    if args.export:
        data = get_companies(city=args.city, country=args.country)
        if data:
            export_to_excel(data, args.query, args.city, args.country)
        return

    run_id        = str(uuid.uuid4())[:8]
    clean_query   = args.query.strip().title()
    clean_city    = args.city.strip().title()
    clean_country = args.country.strip().title()

    results = await scrape_google_maps(
        clean_query, clean_city, clean_country,
        args.start, args.end, run_id
    )

    if not results:
        print("\n❌ No companies found.")
        return

    inserted = updated = skipped = 0
    for company in results:
        status = upsert_company(run_id, company)
        if status == "inserted": inserted += 1
        elif status == "updated": updated  += 1
        else:                     skipped  += 1

    save_search(run_id, args.query, args.city, args.country, args.start, args.end, len(results))
    export_to_excel(results, args.query, args.city, args.country)

    print(f"\n📊 Summary — Scraped: {len(results)} | Inserted: {inserted} | Updated: {updated} | Skipped: {skipped}")


if __name__ == "__main__":
    asyncio.run(main())
