"""
ReachCT — api.py  v2.1
FastAPI backend — scraper + auth + user databases + Gemini upload cleaning.
"""

import os, sys, uuid, json, asyncio, io, re
from dotenv import load_dotenv
load_dotenv()
from datetime import datetime
from typing import Optional, List

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from fastapi import FastAPI, HTTPException, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from database import (
    init_db, save_search, upsert_company, get_companies,
    upsert_user, get_user_by_email, get_filters,
    create_user_database, get_user_databases, get_user_database,
    delete_user_database, get_db_entries, add_db_entries,
    update_db_entry, delete_db_entry, rename_column_in_db,
    add_collaborator, get_collaborators, remove_collaborator,
)
from reachct import scrape_google_maps, export_to_excel
from auth    import verify_google_token, create_jwt, decode_jwt

# ── Inline AI helpers (Claude Haiku) ─────────────────────────────────────────
import anthropic as _anthropic
import re as _re

def _get_claude():
    return _anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY",""))

def _generate(prompt: str, max_tokens: int = 2000) -> str:
    msg = _get_claude().messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=max_tokens,
        messages=[{"role":"user","content":prompt}]
    )
    return msg.content[0].text.strip()

def _generate_json(prompt: str, max_tokens: int = 2000) -> str:
    text = _generate(prompt, max_tokens)
    text = _re.sub(r"```json\s*","",text)
    text = _re.sub(r"```\s*","",text)
    return text.strip()


app = FastAPI(title="ReachCT API", version="2.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Queue ─────────────────────────────────────────────────────────────────────
import threading, queue as queue_module
jobs: dict   = {}
search_queue = queue_module.Queue()
queue_lock   = threading.Lock()

def queue_worker():
    while True:
        try:
            job_id, query, city, country, start, end = search_queue.get(timeout=300)
            jobs[job_id]["status"]         = "starting"
            jobs[job_id]["queue_position"] = 0
            for idx, j in enumerate([j for j in jobs.values() if j["status"] == "queued"]):
                j["queue_position"] = idx + 1
            t = threading.Thread(target=run_scrape_job_thread, args=(job_id, query, city, country, start, end))
            t.start(); t.join()
            search_queue.task_done()
        except queue_module.Empty: continue
        except Exception as e: print(f"❌ Queue worker error: {e}"); continue

threading.Thread(target=queue_worker, daemon=True).start()

# ── Auth helpers ──────────────────────────────────────────────────────────────
def get_current_user(authorization: str = None) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_jwt(authorization[7:])
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return payload

# ── Startup ───────────────────────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    init_db()
    try:
        from database import init_linkedin_table
        init_linkedin_table()
    except Exception as e:
        print(f"⚠️  LinkedIn table init: {e}")
    print("✅ ReachCT API ready")

@app.get("/api/health")
def health():
    return {"status": "ok", "time": datetime.now().isoformat()}

# ── Auth ──────────────────────────────────────────────────────────────────────
class GoogleAuthRequest(BaseModel):
    credential: str

@app.post("/api/auth/google")
def auth_google(body: GoogleAuthRequest):
    try:
        info = verify_google_token(body.credential)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid Google token")
    user  = upsert_user(info["sub"], info.get("email",""), info.get("name",""), info.get("picture",""))
    token = create_jwt(user["id"], info.get("email",""), info.get("name",""), info.get("picture",""))
    return {"token": token, "user": {"id": user["id"], "email": info.get("email",""), "name": info.get("name",""), "picture": info.get("picture","")}}

@app.get("/api/auth/me")
def auth_me(authorization: str = Header(default=None)):
    payload = get_current_user(authorization)
    return {"id": payload["sub"], "email": payload["email"], "name": payload["name"], "picture": payload["picture"]}

# ── Scrape ────────────────────────────────────────────────────────────────────
@app.get("/api/scrape")
async def start_scrape(query: str, city: str, country: str, start: int = 0, end: int = 25):
    query = query.strip(); city = city.strip().title(); country = country.strip().title()
    if not query or not city or not country:
        raise HTTPException(status_code=400, detail="query, city and country are required")
    if end <= start: raise HTTPException(status_code=400, detail="end must be greater than start")
    if (end - start) > 50: raise HTTPException(status_code=400, detail="Maximum 50 listings per search.")
    job_id = str(uuid.uuid4())[:8]
    with queue_lock:
        qr = sum(1 for j in jobs.values() if j["status"] in ("running","queued"))
        jobs[job_id] = {"status":"queued","queue_position":qr,"progress":0,"total":end-start,
            "total_on_maps":None,"processing":None,"results":[],"error":None,
            "query":query,"city":city,"country":country}
        search_queue.put((job_id, query, city, country, start, end))
    return {"job_id": job_id, "message": "Scrape started" if qr==0 else f"Queued at position {qr}", "queue_position": qr}

def run_scrape_job_thread(job_id, query, city, country, start, end):
    loop = asyncio.new_event_loop(); asyncio.set_event_loop(loop)
    try: loop.run_until_complete(run_scrape_job(job_id, query, city, country, start, end))
    finally: loop.close()

async def run_scrape_job(job_id, query, city, country, start, end):
    try:
        jobs[job_id]["status"] = "running"
        results = await scrape_google_maps(query, city, country, start, end, job_id, jobs=jobs, job_id=job_id)
        for c in results: upsert_company(job_id, c)
        save_search(job_id, query, city, country, start, end, len(results))
        jobs[job_id]["status"]  = "cancelled" if jobs[job_id].get("status") == "cancelling" else "done"
        jobs[job_id]["results"] = results
    except Exception as e:
        jobs[job_id]["status"] = "error"; jobs[job_id]["error"] = str(e)

@app.post("/api/job/{job_id}/cancel")
def cancel_job(job_id: str):
    job = jobs.get(job_id)
    if not job: raise HTTPException(status_code=404, detail="Job not found")
    if job["status"] in ("running","queued"): job["status"] = "cancelling"; return {"message":"Cancellation requested"}
    return {"message": f"Job already {job['status']}"}

@app.get("/api/job/{job_id}")
def get_job(job_id: str):
    job = jobs.get(job_id)
    if not job: raise HTTPException(status_code=404, detail="Job not found")
    return job

# ── Companies ─────────────────────────────────────────────────────────────────
class MultiFilterRequest(BaseModel):
    queries:   List[str] = []
    cities:    List[str] = []
    countries: List[str] = []

@app.post("/api/companies/multi")
def get_companies_multi(body: MultiFilterRequest):
    data = get_companies(queries=body.queries, cities=body.cities, countries=body.countries)
    return {"companies": data, "total": len(data)}

@app.get("/api/companies")
def get_all_companies(city: Optional[str]=None, country: Optional[str]=None, query: Optional[str]=None):
    if city: city = city.strip().title()
    if country: country = country.strip().title()
    if query: query = query.strip()
    data = get_companies(query=query, city=city, country=country)
    return {"companies": data, "total": len(data)}

@app.get("/api/filters")
def get_filters_endpoint():
    return get_filters()

@app.get("/api/export")
def export_shared(query: str="", city: str="", country: str=""):
    data = get_companies(city=city.strip().title(), country=country.strip().title())
    if not data: raise HTTPException(status_code=404, detail="No companies found")
    filename = export_to_excel(data, query or "export", city, country)
    if not filename or not os.path.exists(filename):
        raise HTTPException(status_code=500, detail="Failed to generate Excel file")
    return FileResponse(path=filename, filename=os.path.basename(filename),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/searches")
def get_searches_endpoint():
    from database import get_searches
    return {"searches": get_searches()}

# ── User Databases ────────────────────────────────────────────────────────────
class CreateDBRequest(BaseModel):
    name: str
    kind: str = "maps"   # "maps" or "linkedin"

@app.post("/api/databases")
def create_db(body: CreateDBRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    return create_user_database(int(user["sub"]), body.name.strip(), body.kind)

@app.get("/api/databases")
def list_dbs(authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    return get_user_databases(int(user["sub"]))

@app.delete("/api/databases/{db_id}")
def delete_db(db_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    if not delete_user_database(db_id, int(user["sub"])):
        raise HTTPException(status_code=404, detail="Database not found or not owner")
    return {"deleted": True}

# ── Database entries ──────────────────────────────────────────────────────────
@app.get("/api/databases/{db_id}/entries")
def get_entries(db_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    if not get_user_database(db_id, int(user["sub"])):
        raise HTTPException(status_code=403, detail="Access denied")
    return get_db_entries(db_id)

class AddEntriesRequest(BaseModel):
    rows: List[dict]

@app.post("/api/databases/{db_id}/entries")
def add_entries(db_id: int, body: AddEntriesRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied")
    if db.get("role") == "viewer": raise HTTPException(status_code=403, detail="Viewers cannot add entries")
    return add_db_entries(db_id, body.rows)

class UpdateEntryRequest(BaseModel):
    data: dict

@app.patch("/api/databases/{db_id}/entries/{entry_id}")
def update_entry(db_id: int, entry_id: int, body: UpdateEntryRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied")
    if db.get("role") == "viewer": raise HTTPException(status_code=403, detail="Viewers cannot edit")
    row = update_db_entry(entry_id, db_id, body.data)
    if not row: raise HTTPException(status_code=404, detail="Entry not found")
    return row

@app.delete("/api/databases/{db_id}/entries/{entry_id}")
def delete_entry(db_id: int, entry_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied")
    if db.get("role") == "viewer": raise HTTPException(status_code=403, detail="Viewers cannot delete")
    if not delete_db_entry(entry_id, db_id):
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"deleted": True}

# ── Rename column ─────────────────────────────────────────────────────────────
class RenameColRequest(BaseModel):
    old_name: str
    new_name: str

@app.post("/api/databases/{db_id}/rename-column")
def rename_column(db_id: int, body: RenameColRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied")
    if db.get("role") == "viewer": raise HTTPException(status_code=403, detail="Viewers cannot rename columns")
    count = rename_column_in_db(db_id, body.old_name, body.new_name)
    return {"renamed": count}

# ── Export user database to Excel ─────────────────────────────────────────────
@app.get("/api/databases/{db_id}/export")
def export_user_db(db_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied")
    entries = get_db_entries(db_id)
    if not entries: raise HTTPException(status_code=404, detail="No entries to export")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    # Derive columns
    cols = []
    seen = set()
    for e in entries:
        for k in (e.get("data") or {}).keys():
            if k not in seen: seen.add(k); cols.append(k)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = db["name"][:31]

    hdr_fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
    hdr_font = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill; cell.font = hdr_font
        ws.column_dimensions[cell.column_letter].width = 25

    for ri, entry in enumerate(entries, 2):
        data = entry.get("data") or {}
        for ci, col in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=data.get(col, ""))

    stamp    = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"reachct_{db['name'].replace(' ','_')}_{stamp}.xlsx"
    wb.save(filename)
    return FileResponse(path=filename, filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Upload with Claude cleaning ───────────────────────────────────────────────
@app.post("/api/databases/{db_id}/upload")
async def upload_file(db_id: int, file: UploadFile = File(...), authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied")
    if db.get("role") == "viewer": raise HTTPException(status_code=403, detail="Viewers cannot upload")

    try:
        import pandas as pd

        contents = await file.read()
        buf      = io.BytesIO(contents)

        if file.filename.lower().endswith(".csv"):
            df = pd.read_csv(buf, dtype=str, header=None)
        else:
            df = pd.read_excel(buf, dtype=str, header=None)

        # Remove completely empty rows/cols
        df = df.dropna(how="all").reset_index(drop=True)
        df = df.dropna(axis=1, how="all")

        # Convert to raw text for Claude to analyze
        raw_text = df.to_csv(index=False, header=False)

        # Ask Gemini to identify and standardize the data
        _prompt = f"""You are a data extraction specialist. Your job is to find company contact information in messy, unlabeled spreadsheet data.

The data below may have: no headers, wrong column order, empty rows, mixed languages, extra noise, or partial information.

Raw data:
{raw_text}

Instructions:
1. Scan every non-empty row for company information
2. Use context clues to identify: company name (usually a proper noun/business name), email (contains @), phone (digits with +/spaces/dashes), website (contains . or http), city (place name), country (country name), company type (business category)
3. Even if a row is missing most fields, include it if it has at least a company name OR email
4. Return ONLY a valid JSON array — no explanation, no markdown, no code fences
5. Each object must have exactly these keys (empty string "" if unknown): name, email, phone, website, city, country, company_type
6. Clean: lowercase emails, keep only digits/+/spaces in phones, skip rows that are completely empty or clearly not company data

Return ONLY the JSON array starting with [ and ending with ]"""
        _ai_result = _generate_json(_prompt, max_tokens=4000)

        response_text = _ai_result
        # Strip markdown if present
        response_text = re.sub(r"```json\s*", "", response_text)
        response_text = re.sub(r"```\s*", "", response_text)

        cleaned_rows = json.loads(response_text)

        if not isinstance(cleaned_rows, list):
            raise ValueError("Gemini did not return a list")

        # Save to user database
        entries = add_db_entries(db_id, [dict(row) for row in cleaned_rows])

        # Save to shared companies table
        for row in cleaned_rows:
            if row.get("name"):
                upsert_company("upload", {
                    "name":         row.get("name",""),
                    "email":        row.get("email",""),
                    "phone":        row.get("phone",""),
                    "website":      row.get("website",""),
                    "city":         row.get("city",""),
                    "country":      row.get("country",""),
                    "company_type": row.get("company_type",""),
                    "maps_url":     "",
                })

        cols = ["name","email","phone","website","city","country","company_type"]
        return {"inserted": len(entries), "columns": cols, "cleaned_by": "gemini"}

    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="Gemini could not parse the file. Make sure it contains company data.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Upload failed: {str(e)}")

# ── Pull from shared DB into user database ────────────────────────────────────
class PullToDBRequest(BaseModel):
    queries:   List[str] = []
    cities:    List[str] = []
    countries: List[str] = []

@app.post("/api/databases/{db_id}/pull")
def pull_to_db(db_id: int, body: PullToDBRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied")
    if db.get("role") == "viewer": raise HTTPException(status_code=403, detail="Viewers cannot pull data")
    companies = get_companies(queries=body.queries, cities=body.cities, countries=body.countries)
    if not companies: return {"inserted": 0, "message": "No companies found matching filters"}
    rows = [{"company_id": str(c.get("id","")), "name": c.get("name",""), "email": c.get("email",""),
             "phone": c.get("phone",""), "website": c.get("website",""), "city": c.get("city",""),
             "country": c.get("country",""), "company_type": c.get("company_type","")} for c in companies]
    entries = add_db_entries(db_id, rows)
    return {"inserted": len(entries), "columns": ["name","email","phone","website","city","country","company_type"]}

# ── Add specific rows from search/pull to a user database ─────────────────────
class AddRowsToDBRequest(BaseModel):
    db_id: int
    rows:  List[dict]

@app.post("/api/databases/add-rows")
def add_rows_to_db(body: AddRowsToDBRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    db   = get_user_database(body.db_id, int(user["sub"]))
    if not db: raise HTTPException(status_code=403, detail="Access denied or database not found")
    if db.get("role") == "viewer": raise HTTPException(status_code=403, detail="Viewers cannot add entries")
    entries = add_db_entries(body.db_id, body.rows)
    return {"inserted": len(entries)}

# ── Collaborators ─────────────────────────────────────────────────────────────
class AddCollaboratorRequest(BaseModel):
    email: str
    role:  str = "viewer"

@app.post("/api/databases/{db_id}/collaborators")
def add_collab(db_id: int, body: AddCollaboratorRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    try: return add_collaborator(db_id, int(user["sub"]), body.email, body.role)
    except ValueError as e: raise HTTPException(status_code=400, detail=str(e))
    except PermissionError as e: raise HTTPException(status_code=403, detail=str(e))

@app.get("/api/databases/{db_id}/collaborators")
def list_collabs(db_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    if not get_user_database(db_id, int(user["sub"])): raise HTTPException(status_code=403, detail="Access denied")
    return get_collaborators(db_id)

@app.delete("/api/databases/{db_id}/collaborators/{target_user_id}")
def remove_collab(db_id: int, target_user_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    try:
        if not remove_collaborator(db_id, int(user["sub"]), target_user_id):
            raise HTTPException(status_code=404, detail="Collaborator not found")
        return {"deleted": True}
    except PermissionError as e: raise HTTPException(status_code=403, detail=str(e))

# ── Admin ─────────────────────────────────────────────────────────────────────
@app.get("/api/admin/jobs")
def admin_get_jobs():
    return {"jobs": [{"id": jid, **job} for jid, job in jobs.items()]}

@app.post("/api/admin/cancel-all")
def admin_cancel_all():
    cancelled = []
    for jid, job in jobs.items():
        if job["status"] in ("running","queued","starting"):
            job["status"] = "cancelling"; cancelled.append(jid)
    return {"cancelled": cancelled, "count": len(cancelled)}

# ── Shared DB upload (Push tab on DatabasePage, no user DB needed) ────────────
@app.post("/api/upload-shared")
async def upload_shared(file: UploadFile = File(...), authorization: str = Header(default=None)):
    """Upload Excel/CSV directly to the shared companies table via Claude cleaning."""
    user = get_current_user(authorization)  # must be logged in
    try:
        import pandas as pd

        contents = await file.read()
        buf      = io.BytesIO(contents)
        if file.filename.lower().endswith(".csv"):
            df = pd.read_csv(buf, dtype=str, header=None)
        else:
            df = pd.read_excel(buf, dtype=str, header=None)

        df = df.dropna(how="all").reset_index(drop=True)
        df = df.dropna(axis=1, how="all")
        raw_text = df.to_csv(index=False, header=False)

        _prompt = f"""You are a data cleaning assistant. Analyze this raw spreadsheet data and extract company contact information.

Raw data:
{raw_text}

Return ONLY a valid JSON array. Each object must have these keys (use "" if unknown):
name, email, phone, website, city, country, company_type

Rules: skip empty rows, clean phones (digits/+ only), lowercase emails, don't invent data.
Return ONLY the JSON array, no explanation."""
        _ai_result = _generate_json(_prompt, max_tokens=4000)

        response_text = re.sub(r"```json\s*","",_ai_result)
        response_text = re.sub(r"```\s*","",response_text)
        cleaned_rows  = json.loads(response_text)

        inserted = 0
        for row in cleaned_rows:
            if row.get("name"):
                upsert_company("upload_shared", {
                    "name":         row.get("name",""),
                    "email":        row.get("email",""),
                    "phone":        row.get("phone",""),
                    "website":      row.get("website",""),
                    "city":         row.get("city",""),
                    "country":      row.get("country",""),
                    "company_type": row.get("company_type",""),
                    "maps_url":     "",
                })
                inserted += 1

        return {"inserted": inserted, "cleaned_by": "gemini"}

    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="Gemini could not parse the file. Please ensure it contains company data.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Upload failed: {str(e)}")


# ── ReachAI — Claude agent endpoint ──────────────────────────────────────────
from ai_tools import (
    tool_list_databases, tool_get_database_contents, tool_get_database_stats,
    tool_pull_from_database, tool_save_to_database, tool_create_database,
    tool_search_google_maps,
)

REACHAI_TOOLS = [
    {
        "name": "list_databases",
        "description": "List all databases the user has access to, with row counts.",
        "input_schema": {"type": "object", "properties": {}, "required": []}
    },
    {
        "name": "get_database_contents",
        "description": "Get the entries/rows from a specific user database.",
        "input_schema": {
            "type": "object",
            "properties": {
                "db_id": {"type": "integer", "description": "The database ID"}
            },
            "required": ["db_id"]
        }
    },
    {
        "name": "get_database_stats",
        "description": "Get stats about the shared ReachCT database — total companies, breakdown by country, email find rate.",
        "input_schema": {"type": "object", "properties": {}, "required": []}
    },
    {
        "name": "pull_from_database",
        "description": "Pull companies from the shared database with optional filters.",
        "input_schema": {
            "type": "object",
            "properties": {
                "queries":   {"type": "array", "items": {"type": "string"}, "description": "Company types e.g. ['Marketing Agency']"},
                "cities":    {"type": "array", "items": {"type": "string"}, "description": "Cities e.g. ['Madrid', 'Berlin']"},
                "countries": {"type": "array", "items": {"type": "string"}, "description": "Countries e.g. ['Spain']"},
            },
            "required": []
        }
    },
    {
        "name": "save_to_database",
        "description": "Save a list of company rows to a user database.",
        "input_schema": {
            "type": "object",
            "properties": {
                "db_id": {"type": "integer", "description": "The database ID to save to"},
                "rows":  {"type": "array",   "items": {"type": "object"}, "description": "List of company objects"}
            },
            "required": ["db_id", "rows"]
        }
    },
    {
        "name": "create_database",
        "description": "Create a new user database with a given name. Only call this after user confirms.",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "The name for the new database"}
            },
            "required": ["name"]
        }
    },
    {
        "name": "search_google_maps",
        "description": "Search Google Maps for companies and scrape their contact info. This runs in the background and may take several minutes.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query":   {"type": "string",  "description": "Business type e.g. 'Marketing Agency'"},
                "city":    {"type": "string",  "description": "City e.g. 'Madrid'"},
                "country": {"type": "string",  "description": "Country e.g. 'Spain'"},
                "start":   {"type": "integer", "description": "Start index (default 0)"},
                "end":     {"type": "integer", "description": "End index max start+50 (default 25)"}
            },
            "required": ["query", "city", "country"]
        }
    },
]

REACHAI_SYSTEM = """You are ReachAI, an intelligent assistant built into ReachCT — a B2B contact intelligence platform.

You have access to tools that let you:
- Search Google Maps for company contacts (takes a few minutes, runs in background)
- Pull companies from the shared ReachCT database
- View and save to the user's personal databases
- Get database statistics

Guidelines:
- Be concise and action-oriented. When a task is done, confirm it and provide the relevant link or data.
- For database links use: /dashboard/db/{id}
- When the user asks you to save to a database by name, first call list_databases to find the right one. If it doesn't exist, ask the user to confirm before creating it.
- When creating a new database for the user, suggest relevant columns based on the data (e.g. Company Name, Email, Phone, Website, City, Country, Company Type) or ask what columns they want.
- Before saving to a database, check its existing columns using list_databases or get_database_contents and match the data structure to those columns.
- When running a search, warn the user it takes a few minutes and report results when done.
- You can answer questions about the data using get_database_stats or get_database_contents.
- Keep responses short unless the user asks for detail.
- When saving search results to a database, only save: name, email, phone, website, city, country, company_type. Never include run_id, maps_url, or other internal fields.
- If a database is empty (0 rows), suggest what columns to add before saving data to it.
"""

class ReachAIRequest(BaseModel):
    messages: list  # full conversation history [{role, content}]

@app.post("/api/ai/chat")
async def reachai_chat(body: ReachAIRequest, authorization: str = Header(default=None)):
    """ReachAI — agentic Claude endpoint with tool use."""
    import threading

    user    = get_current_user(authorization)
    user_id = int(user["sub"])

    messages = list(body.messages)
    MAX_ITERS = 10

    import anthropic as _anthropic
    _client = _anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY",""))

    for _ in range(MAX_ITERS):
        response = _client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            system=REACHAI_SYSTEM,
            tools=REACHAI_TOOLS,
            messages=messages[-8:],  # limit history to last 8 messages
        )

        assistant_content = response.content
        messages.append({"role": "assistant", "content": assistant_content})

        if response.stop_reason == "end_turn":
            text = " ".join(b.text for b in assistant_content if hasattr(b, "text"))
            return {"reply": text, "messages": messages}

        if response.stop_reason != "tool_use":
            text = " ".join(b.text for b in assistant_content if hasattr(b, "text"))
            return {"reply": text, "messages": messages}

        tool_results = []
        for block in assistant_content:
            if block.type != "tool_use":
                continue
            tool_name  = block.name
            tool_input = block.input
            try:
                if tool_name == "list_databases":
                    result = tool_list_databases(user_id)
                elif tool_name == "get_database_contents":
                    result = tool_get_database_contents(user_id, tool_input["db_id"])
                elif tool_name == "get_database_stats":
                    result = tool_get_database_stats(user_id)
                elif tool_name == "pull_from_database":
                    result = tool_pull_from_database(
                        tool_input.get("queries", []),
                        tool_input.get("cities", []),
                        tool_input.get("countries", []),
                    )
                elif tool_name == "save_to_database":
                    result = tool_save_to_database(user_id, tool_input["db_id"], tool_input["rows"])
                elif tool_name == "create_database":
                    result = tool_create_database(user_id, tool_input["name"])
                elif tool_name == "search_google_maps":
                    search_result = {}
                    def run_search():
                        search_result.update(tool_search_google_maps(
                            tool_input["query"], tool_input["city"], tool_input["country"],
                            tool_input.get("start", 0), tool_input.get("end", 25),
                            jobs, search_queue,
                        ))
                    t = threading.Thread(target=run_search)
                    t.start(); t.join()
                    result = search_result
                else:
                    result = {"error": f"Unknown tool: {tool_name}"}
            except Exception as e:
                result = {"error": str(e)}

            tool_results.append({
                "type":        "tool_result",
                "tool_use_id": block.id,
                "content":     json.dumps(result),
            })

        messages.append({"role": "user", "content": tool_results})

    return {"reply": "I reached the maximum number of steps. Please try a simpler request.", "messages": messages}


# ── Mail Campaigns ────────────────────────────────────────────────────────────
from mailrelay import validate_api_key, get_senders, create_group, add_subscribers, create_campaign
from database  import (save_mailrelay_key, get_mailrelay_key,
                        create_campaign_record, get_user_campaigns, delete_campaign_record,
                        init_campaigns_tables)

# Init campaign tables on startup (called after init_db)
try:
    init_campaigns_tables()
except Exception as e:
    print(f"⚠️  Campaign table init: {e}")


class MailrelayKeyRequest(BaseModel):
    api_key: str

@app.post("/api/mailrelay/connect")
def connect_mailrelay(body: MailrelayKeyRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    if not validate_api_key(body.api_key):
        raise HTTPException(status_code=400, detail="Invalid Mailrelay API key — please check and try again")
    save_mailrelay_key(int(user["sub"]), body.api_key)
    return {"connected": True}

@app.get("/api/mailrelay/status")
def mailrelay_status(authorization: str = Header(default=None)):
    user    = get_current_user(authorization)
    api_key = get_mailrelay_key(int(user["sub"]))
    return {"connected": bool(api_key)}

@app.delete("/api/mailrelay/disconnect")
def disconnect_mailrelay(authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    save_mailrelay_key(int(user["sub"]), "")
    return {"disconnected": True}

@app.get("/api/mailrelay/senders")
def get_mailrelay_senders(authorization: str = Header(default=None)):
    user    = get_current_user(authorization)
    api_key = get_mailrelay_key(int(user["sub"]))
    if not api_key:
        raise HTTPException(status_code=400, detail="No Mailrelay account connected")
    return get_senders(api_key)


class GroupData(BaseModel):
    name:   str
    emails: List[str]

class CreateCampaignRequest(BaseModel):
    name:      str
    subject:   str
    body:      str
    contacts:  List[dict]   # flat list for backward compat
    groups:    List[GroupData] = []  # named groups from Step 1
    sender_id: int

@app.post("/api/campaigns")
def create_new_campaign(body: CreateCampaignRequest, authorization: str = Header(default=None)):
    user    = get_current_user(authorization)
    user_id = int(user["sub"])
    api_key = get_mailrelay_key(user_id)
    if not api_key:
        raise HTTPException(status_code=400, detail="No Mailrelay account connected")

    try:
        all_group_ids    = []
        total_subscribed = 0
        total_failed     = 0

        print(f"🔍 Received groups: {body.groups}")
        print(f"🔍 Received contacts: {len(body.contacts)}")
        # Use named groups if provided, otherwise fall back to flat contacts
        groups_to_create = body.groups if body.groups else [
            GroupData(name=body.name, emails=[
                c.get("email","").strip().lower()
                for c in body.contacts
                if c.get("email") and "@" in c.get("email","")
            ])
        ]

        if not groups_to_create or all(len(g.emails)==0 for g in groups_to_create):
            raise HTTPException(status_code=400, detail="No valid email addresses")

        for grp in groups_to_create:
            valid_emails = [e.strip().lower() for e in grp.emails if e and "@" in e]
            if not valid_emails:
                continue

            # 1. Create group with user's chosen name
            group_resp = create_group(api_key, grp.name)
            print(f"🔍 Group response: {group_resp}")
            group_id = (group_resp.get("id") or
                       group_resp.get("data",{}).get("id") or
                       group_resp.get("group",{}).get("id"))

            if not group_id:
                print(f"⚠️ Could not get group_id for {grp.name}")
                continue

            all_group_ids.append(int(group_id))

            # 2. Add subscribers to this group
            print(f"🔍 Adding {len(valid_emails)} emails to group '{grp.name}' (id={group_id})")
            sub_results = add_subscribers(api_key, group_id, valid_emails)
            print(f"🔍 Result: success={sub_results['success']} failed={sub_results['failed']}")
            total_subscribed += sub_results["success"]
            total_failed     += sub_results["failed"]

        if not all_group_ids:
            raise Exception("No groups were created successfully")

        # 3. Create campaign draft — use subject as the campaign name
        campaign_name = body.subject or body.name or "ReachCT Campaign"
        print(f"🔍 Creating campaign '{campaign_name}' with group_ids={all_group_ids} sender_id={body.sender_id}")
        campaign = create_campaign(
            api_key, campaign_name, body.subject,
            body.body or "<p>Email body — edit in Mailrelay before sending.</p>",
            all_group_ids[0], body.sender_id
        )
        print(f"🔍 Campaign response: {campaign}")
        campaign_id = campaign.get("id") or campaign.get("data",{}).get("id") or 0

        # 4. Save to ReachCT DB
        record = create_campaign_record(
            user_id, body.name, body.subject, body.body or "",
            all_group_ids[0], campaign_id, total_subscribed
        )

        return {
            "campaign":           record,
            "subscribers_added":  total_subscribed,
            "subscribers_failed": total_failed,
            "group_ids":          all_group_ids,
            "mailrelay_url":      "https://spain-internship.ipzmarketing.com/admin/campaigns",
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Campaign creation failed: {str(e)}")


@app.get("/api/campaigns")
def list_campaigns_endpoint(authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    return get_user_campaigns(int(user["sub"]))

@app.delete("/api/campaigns/{campaign_id}")
def delete_campaign(campaign_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    if not delete_campaign_record(campaign_id, int(user["sub"])):
        raise HTTPException(status_code=404, detail="Campaign not found")
    return {"deleted": True}


class GenerateCampaignRequest(BaseModel):
    campaign_name: str
    company_type:  str = ""
    sample_contacts: List[dict] = []

@app.post("/api/campaigns/generate")
def generate_campaign_content(body: GenerateCampaignRequest, authorization: str = Header(default=None)):
    """Use Claude to generate email subject and body for a campaign."""
    user   = get_current_user(authorization)

    sample_names = [c.get("name","") for c in body.sample_contacts[:5] if c.get("name")]
    context      = f"Campaign: {body.campaign_name}"
    if body.company_type:
        context += f"\nTarget company type: {body.company_type}"
    if sample_names:
        context += f"\nExample companies: {', '.join(sample_names)}"

        _prompt = f"""Generate a professional B2B outreach email for the following campaign.

{context}

Return ONLY a JSON object with exactly two keys:
- "subject": a compelling email subject line (max 60 chars)
- "body": the email body in HTML format, professional and concise (150-200 words). Use {{{{name}}}} for the contact name and {{{{company}}}} for the company name as merge tags.

Return ONLY the JSON, no explanation."""
        _ai_result = _generate_json(_prompt, max_tokens=800)

    try:
        text = _ai_result
        text = re.sub(r"```json\s*","",text)
        text = re.sub(r"```\s*","",text)
        result = json.loads(text)
        return {"subject": result.get("subject",""), "body": result.get("body","")}
    except:
        raise HTTPException(status_code=500, detail="Failed to generate campaign content")


# ── Email Templates ───────────────────────────────────────────────────────────
from database import (create_template, get_user_templates,
                      update_template, delete_template, init_templates_table)

try:
    init_templates_table()
except Exception as e:
    print(f"⚠️  Template table init: {e}")


class TemplateRequest(BaseModel):
    name:    str
    subject: str = ""
    body:    str = ""

@app.post("/api/templates")
def create_email_template(body: TemplateRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    return create_template(int(user["sub"]), body.name, body.subject, body.body)

@app.get("/api/templates")
def list_templates(authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    return get_user_templates(int(user["sub"]))

@app.patch("/api/templates/{template_id}")
def update_email_template(template_id: int, body: TemplateRequest, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    t    = update_template(template_id, int(user["sub"]), body.name, body.subject, body.body)
    if not t: raise HTTPException(status_code=404, detail="Template not found")
    return t

@app.delete("/api/templates/{template_id}")
def delete_email_template(template_id: int, authorization: str = Header(default=None)):
    user = get_current_user(authorization)
    if not delete_template(template_id, int(user["sub"])):
        raise HTTPException(status_code=404, detail="Template not found")
    return {"deleted": True}


# ── ReachAI email enhancement ─────────────────────────────────────────────────
class EnhanceEmailRequest(BaseModel):
    current_body: str
    instruction:  str
    subject:      str = ""

@app.post("/api/campaigns/enhance")
def enhance_email(body: EnhanceEmailRequest, authorization: str = Header(default=None)):
    """Gemini enhances/appends to the current email body based on user instruction."""
    get_current_user(authorization)

    result = _generate(f"""You are an email writing assistant for a B2B internship placement company.

Current email body:
{body.current_body or "(empty)"}

Subject: {body.subject or "(none)"}

User instruction: {body.instruction}

Task: Follow the user's instruction exactly. You may:
- Write a complete new email if the body is empty
- Enhance, translate, reformat, or add to the existing email
- Add a section, signature, or specific content

Return ONLY the updated/new email body as clean HTML. No explanation. No markdown fences.
Preserve any existing content unless told to replace it.
Add your contribution clearly separated if appending.""", max_tokens=1500)

    return {"body": result}




# ── LinkedIn / People Search ──────────────────────────────────────────────────
from database import (init_linkedin_table, upsert_linkedin_contact,
                      get_linkedin_contacts, get_linkedin_filters)

linkedin_jobs: dict = {}

class LinkedInSearchRequest(BaseModel):
    role:        str = ""
    company:     str = ""
    location:    str = ""
    keyword:     str = ""
    domain:      str = ""   # company domain for email guessing
    max_results: int = 15

@app.post("/api/linkedin/search")
def start_linkedin_search(body: LinkedInSearchRequest, authorization: str = Header(default=None)):
    get_current_user(authorization)
    if not (body.role or body.company or body.keyword):
        raise HTTPException(status_code=400, detail="Provide at least a role, company, or keyword")

    job_id = str(uuid.uuid4())[:8]
    linkedin_jobs[job_id] = {"status":"running","found":0,"results":[],"error":None}

    def run():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            from linkedin import scrape_linkedin_people
            results = loop.run_until_complete(scrape_linkedin_people(
                body.role, body.company, body.location, body.keyword,
                body.domain, min(body.max_results, 30), linkedin_jobs, job_id
            ))
            # Save to shared LinkedIn DB
            for person in results:
                if person.get("linkedin_url"):
                    upsert_linkedin_contact(person)
            linkedin_jobs[job_id]["results"] = results
            linkedin_jobs[job_id]["status"]  = "done"
        except Exception as e:
            linkedin_jobs[job_id]["status"] = "error"
            linkedin_jobs[job_id]["error"]  = str(e)
            print(f"❌ LinkedIn search error: {e}")
        finally:
            loop.close()

    threading.Thread(target=run, daemon=True).start()
    return {"job_id": job_id}

@app.get("/api/linkedin/status/{job_id}")
def linkedin_status(job_id: str, authorization: str = Header(default=None)):
    get_current_user(authorization)
    job = linkedin_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job

@app.get("/api/linkedin/filters")
def linkedin_filters(authorization: str = Header(default=None)):
    get_current_user(authorization)
    return get_linkedin_filters()

class LinkedInPullRequest(BaseModel):
    job_titles: List[str] = []
    companies:  List[str] = []
    locations:  List[str] = []

@app.post("/api/linkedin/pull")
def linkedin_pull(body: LinkedInPullRequest, authorization: str = Header(default=None)):
    get_current_user(authorization)
    all_results = []
    seen = set()
    # If no filters, pull everything
    combos = []
    if body.job_titles or body.companies or body.locations:
        jt = body.job_titles or [""]
        co = body.companies or [""]
        lo = body.locations or [""]
        for j in jt:
            for c in co:
                for l in lo:
                    combos.append((j, c, l))
    else:
        combos = [("", "", "")]

    for j, c, l in combos:
        for row in get_linkedin_contacts(j, c, l):
            if row["id"] not in seen:
                seen.add(row["id"])
                all_results.append(row)
    return {"results": all_results, "count": len(all_results)}


# ── LinkedIn Bulk Search ──────────────────────────────────────────────────────
class LinkedInBulkRequest(BaseModel):
    items:           List[str] = []   # pasted emails/domains/companies
    from_db_id:      int = 0           # optional: pull companies from a Maps DB
    role:            str = ""
    location:        str = ""
    max_per_company: int = 5

@app.post("/api/linkedin/bulk")
def start_linkedin_bulk(body: LinkedInBulkRequest, authorization: str = Header(default=None)):
    user    = get_current_user(authorization)
    user_id = int(user["sub"])

    raw_items = list(body.items)

    # If pulling from a Maps database, extract company names + domains from website
    if body.from_db_id:
        db = get_user_database(body.from_db_id, user_id)
        if db:
            for entry in get_db_entries(body.from_db_id):
                data = entry.get("data", {})
                website = data.get("website", "")
                email   = data.get("email", "")
                name    = data.get("name", "")
                if website:
                    raw_items.append(website)
                elif email and "@" in email:
                    raw_items.append(email)
                elif name:
                    raw_items.append(name)

    if not raw_items:
        raise HTTPException(status_code=400, detail="No companies/emails/domains provided")

    from linkedin import parse_bulk_input
    targets = parse_bulk_input(raw_items)
    if not targets:
        raise HTTPException(status_code=400, detail="Could not parse any valid targets")

    job_id = str(uuid.uuid4())[:8]
    linkedin_jobs[job_id] = {"status":"running","found":0,"results":[],"error":None,
                             "processing":None,"company_index":0,"total_companies":len(targets)}

    def run():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            from linkedin import scrape_linkedin_bulk
            results = loop.run_until_complete(scrape_linkedin_bulk(
                targets, body.role, body.location,
                min(body.max_per_company, 10), linkedin_jobs, job_id
            ))
            for person in results:
                if person.get("linkedin_url"):
                    upsert_linkedin_contact(person)
            linkedin_jobs[job_id]["results"] = results
            linkedin_jobs[job_id]["status"]  = "done"
        except Exception as e:
            linkedin_jobs[job_id]["status"] = "error"
            linkedin_jobs[job_id]["error"]  = str(e)
            print(f"❌ LinkedIn bulk error: {e}")
        finally:
            loop.close()

    threading.Thread(target=run, daemon=True).start()
    return {"job_id": job_id, "targets": len(targets)}


# ── LinkedIn Smart Search (uses companies DB) ─────────────────────────────────
class LinkedInSmartRequest(BaseModel):
    company_type: str
    city:         str
    role:         str = "HR"
    start:        int = 0
    end:          int = 25

@app.post("/api/linkedin/smart")
def start_linkedin_smart(body: LinkedInSmartRequest, authorization: str = Header(default=None)):
    """
    Smart LinkedIn search — pulls companies from shared DB by type+city,
    then searches LinkedIn for decision makers at each company.
    """
    get_current_user(authorization)

    # Get companies from shared DB matching type + city
    companies = get_companies(
        queries=[body.company_type],
        cities=[body.city.strip().title()]
    )

    if not companies:
        raise HTTPException(status_code=404,
            detail=f"No companies found for {body.company_type} in {body.city}. Run a Google Maps search first.")

    # Apply start/end range
    total = len(companies)
    companies = companies[body.start:body.end]
    print(f"🔍 Smart LinkedIn: {len(companies)} companies (range {body.start}→{body.end} of {total}) for {body.company_type} in {body.city}")

    # Build targets from DB — company name + domain from website
    targets = []
    for c in companies:
        name    = c.get("name", "")
        website = c.get("website", "")
        email   = c.get("email", "")
        # Extract domain from website or email
        domain = ""
        if website:
            domain = website.lower().replace("https://","").replace("http://","").replace("www.","").strip("/").split("/")[0]
        elif email and "@" in email:
            domain = email.split("@")[-1].strip().lower()
        if name:
            targets.append({"company": name, "domain": domain})

    if not targets:
        raise HTTPException(status_code=404, detail="No valid company targets found")

    job_id = str(uuid.uuid4())[:8]
    linkedin_jobs[job_id] = {
        "status":           "running",
        "found":            0,
        "results":          [],
        "error":            None,
        "processing":       None,
        "company_index":    0,
        "total_companies":  len(targets),
    }

    def run():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            from linkedin import scrape_linkedin_bulk
            results = loop.run_until_complete(scrape_linkedin_bulk(
                targets, body.role, body.city,
                1, linkedin_jobs, job_id  # 1 person per company max
            ))
            for person in results:
                if person.get("linkedin_url"):
                    upsert_linkedin_contact(person)
            linkedin_jobs[job_id]["results"] = results
            linkedin_jobs[job_id]["status"]  = "done"
        except Exception as e:
            linkedin_jobs[job_id]["status"] = "error"
            linkedin_jobs[job_id]["error"]  = str(e)
            print(f"❌ LinkedIn smart search error: {e}")
        finally:
            loop.close()

    threading.Thread(target=run, daemon=True).start()
    return {"job_id": job_id, "total_companies": len(targets)}